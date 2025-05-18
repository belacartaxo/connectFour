import Connect4 as C4
import numpy as np
import openpyxl 


workbook = openpyxl.load_workbook("FILE_NAME.XLSX")
worksheet = workbook.active


# Simulate n games
# save the final board into a 1D array and the winner. 
# I have no clue how to determine moves in the middle, so that is something to think about

# EM SUMA, ONLY STORE THE FINAL BAORD AND THE WINNER, possibly make new functions? 
# Research on Python excel magic
# should be done by today





def run_simulation():
    boardStates, playerTurns, optimalMoves = C4.random_vs_random()   # RETURNS 3 lists, all the baords across the simulated game, player turns and respective optimal move
    
    return boardStates, playerTurns, optimalMoves


def generate_db(iterations=150, last_row=1):  # last_row defualt = 1, if last row written in the dataset is n, set last_row = n
    last_updated_row = 0
    for i in range(iterations):
        print(f"Initiating simulation {i + 1}:")
        boardStates, playerTurns, optimalMoves = run_simulation()

        print("Writting Data...")
        # Since each of the board size varries,
        allocate = len(boardStates)
        for j in range(allocate):
            for k in range(1, 43):
                worksheet.cell(row=j + 1 + last_row + last_updated_row, column=k, value=f"{boardStates[j][k - 1]}")
            worksheet.cell(row=j + 1 + last_row + last_updated_row, column=43, value=f"{playerTurns[j]}")
            worksheet.cell(row=j + 1 + last_row + last_updated_row, column=44, value=optimalMoves[j])
            workbook.save("FILE_NAME.XLSX")

        last_updated_row += allocate




        print(f"simulation {i + 1} has completed!")
        print("Upload completed")
        print(f"Progression { round(((i + 1)/iterations) * 100, 2)}%")
        print("")


generate_db(last_row=8466)

